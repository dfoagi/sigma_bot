from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

LOG_FILE = os.path.join("logs", "logs.xlsx")


def log_to_excel(username, user_id, question, answer, input_tokens, output_tokens, chapter_ids, chapter_scores, model):
    if os.path.exists(LOG_FILE):
        wb = load_workbook(LOG_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Дата", "Время", "Логин", "Telegram ID", "Вопрос", "Ответ", "модель",
                   "Токены вопрос", "Токены ответ", "id документов", "Score"])
        for i, width in enumerate([10, 10, 15, 11, 50, 50, 25, 15, 15, 15, 33], 1):
            ws.column_dimensions[get_column_letter(i)].width = width
    
    datetime_now = datetime.now()
    ws.append([
        datetime_now.strftime("%d.%m.%Y"),
        datetime_now.strftime("%H:%M:%S"),
        username or "—",
        user_id,
        question,
        answer,
        model,
        input_tokens,
        output_tokens,
        chapter_ids,
        chapter_scores,
    ])

    wb.save(LOG_FILE)
